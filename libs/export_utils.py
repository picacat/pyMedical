from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Alignment, Protection, Font
import subprocess

from libs import string_utils
from libs import number_utils
from libs import case_utils
from libs import date_utils
from libs import nhi_utils


light_blue = PatternFill(start_color='99CCFF', end_color='99CCFF', fill_type='solid')
light_yellow = PatternFill(start_color='FFFF99', end_color='FFFF99', fill_type='solid')
light_gray = PatternFill(start_color='C0C0C0', end_color='C0C0C0', fill_type='solid')
salmon = PatternFill(start_color='FFB266', end_color='FFB266', fill_type='solid')
purple = PatternFill(start_color='FF9999', end_color='FF9999', fill_type='solid')
align_center = Alignment(horizontal='center', vertical='center')
bold = Font(bold=True)
side = Side(border_style='thin', color='000000')
border = Border(top=side, bottom=side, left=side, right=side)


def export_table_widget_to_excel(
        excel_file_name, table_widget, hidden_column=None, numeric_cell=None, title=None):
    if numeric_cell is None:
        numeric_cell = []
    wb = Workbook()
    ws = wb.active
    ws.title = 'sheet1'

    header_row = []
    for col_no in range(table_widget.columnCount()):
        if hidden_column is not None and col_no in hidden_column:
            continue

        header_row.append(table_widget.horizontalHeaderItem(col_no).text())

    if title is not None:
        ws.append([title])

    ws.append(header_row)

    for row_no in range(table_widget.rowCount()):
        row = []
        for col_no in range(table_widget.columnCount()):
            if hidden_column is not None and col_no in hidden_column:
                continue

            item = table_widget.item(row_no, col_no)
            if item is not None:
                item_text = item.text()
            else:
                item_text = ''

            if numeric_cell is not None and col_no in numeric_cell:
                item_text = item_text.replace(',', '')
                item_text = number_utils.get_float(item_text)

            row.append(item_text)

        ws.append(row)

    wb.save(excel_file_name)
    subprocess.Popen([excel_file_name], shell=True)


# 匯出日報表 From medical_record_list 2019.07.01 板橋新生堂
def export_daily_medical_records_to_excel(database, system_settings, excel_file_name, table_widget):
    row_range = [
        'A', 'B', 'C', 'D', 'E', 'F', 'G', 'H', 'I', 'J', 'K', 'L', 'M',
        'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U'
    ]

    workbook = Workbook()
    sheet = workbook.active
    sheet.title = '門診日報表資料'

    sheet.append([
        '{clinic_name} 日報表'.format(
            clinic_name=system_settings.field('院所名稱'),
        )
    ])
    sheet.merge_cells('A1:U1')
    row_property = sheet.row_dimensions[1]
    row_property.height = 30
    row_property.alignment = align_center
    row_property.font = bold

    header_row = [
        '日期', '班別', '醫師', '保險', '病歷號', '病患姓名', '診號', '卡序', '療程',
        '掛號費', '門診負擔', '藥品負擔', '自費品項', '自費金額',
        '推拿金額', '推拿自費品項', '推拿自費金額', '推拿師', '押單', '備註', '合計',
    ]
    sheet.append(header_row)
    sheet.row_dimensions[2].height = 30

    sheet.column_dimensions['A'].width = 14
    sheet.column_dimensions['B'].width = 5
    sheet.column_dimensions['C'].width = 10
    sheet.column_dimensions['D'].width = 5
    sheet.column_dimensions['F'].width = 10
    sheet.column_dimensions['G'].width = 6
    sheet.column_dimensions['H'].width = 6
    sheet.column_dimensions['I'].width = 5
    sheet.column_dimensions['K'].width = 10
    sheet.column_dimensions['L'].width = 10
    sheet.column_dimensions['M'].width = 30
    sheet.column_dimensions['N'].width = 10
    sheet.column_dimensions['O'].width = 10
    sheet.column_dimensions['P'].width = 30
    sheet.column_dimensions['Q'].width = 15
    sheet.column_dimensions['R'].width = 10
    sheet.column_dimensions['U'].width = 13

    total_regist_fee = 0
    total_diag_share_fee = 0
    total_drug_share_fee = 0
    total_purchase_item_fee = 0
    total_massage_item_fee = 0
    total_deposit_fee = 0
    ins_count = 0

    daily_regist_fee = 0
    daily_diag_share_fee = 0
    daily_drug_share_fee = 0
    daily_purchase_item_fee = 0
    daily_massage_item_fee = 0
    daily_deposit_fee = 0
    daily_total = 0
    daily_ins_count = 0

    for row_no in range(table_widget.rowCount()):
        check_box = table_widget.cellWidget(row_no, 1)
        if not check_box.isChecked():
            continue

        case_key = table_widget.item(row_no, 0).text()
        sql = '''
            SELECT 
                cases.DepositFee, cases.Massager, patient.DiscountType
            FROM cases
                LEFT JOIN patient ON patient.PatientKey = cases.PatientKey
            WHERE
                CaseKey = {case_key}
        '''.format(
            case_key=case_key
        )
        case_row = database.select_record(sql)
        if len(case_row) > 0:
            case_row = case_row[0]
            deposit_fee = number_utils.get_integer(case_row['DepositFee'])
            massager = string_utils.xstr(case_row['Massager'])
            discount_type = string_utils.xstr(case_row['DiscountType'])
        else:
            deposit_fee = 0
            massager = ''
            discount_type = ''

        case_date = table_widget.item(row_no, 2).text()[:10]
        period = table_widget.item(row_no, 3).text()[:1]
        next_case_date = table_widget.item(row_no+1, 2)
        if next_case_date is not None:
            next_case_date = next_case_date.text()[:10]

        next_period = table_widget.item(row_no+1, 3)
        if next_period is not None:
            next_period = next_period.text()[:1]

        ins_type = table_widget.item(row_no, 12).text()
        if ins_type == '健保':
            ins_count += 1

        card = table_widget.item(row_no, 15).text()
        if card == '不需取得':
            card = ''

        regist_fee = number_utils.get_integer(table_widget.item(row_no, 21).text())
        diag_share_fee = number_utils.get_integer(table_widget.item(row_no, 22).text())
        drug_share_fee = number_utils.get_integer(table_widget.item(row_no, 23).text())

        total_regist_fee += regist_fee
        total_diag_share_fee += diag_share_fee
        total_drug_share_fee += drug_share_fee
        total_deposit_fee += deposit_fee

        sale_item = ''

        row = [
            case_date,
            period,
            table_widget.item(row_no, 17).text(),
            ins_type,
            table_widget.item(row_no, 8).text(),
            table_widget.item(row_no, 9).text(),
            table_widget.item(row_no, 7).text(),
            card,
            table_widget.item(row_no, 16).text(),

            regist_fee,
            diag_share_fee,
            drug_share_fee,
            sale_item,
            number_utils.get_integer(table_widget.item(row_no, 24).text()),
            None, None,
            0,  # massage_fee
            massager,
            deposit_fee,
            discount_type,
        ]
        sql = '''
            SELECT * FROM prescript
            WHERE
                CaseKey = {case_key} AND
                MedicineSet >= 2 AND
                Amount > 0
            ORDER BY PrescriptKey
        '''.format(
            case_key=case_key,
        )
        prescript_rows = database.select_record(sql)
        if len(prescript_rows) <= 0:
            sheet.append(row)
        else:
            for prescript_row_no, prescript_row in zip(range(len(prescript_rows)), prescript_rows):
                if massager == '':
                    pres_days = case_utils.get_pres_days(database, case_key, prescript_row['MedicineSet'])
                    if pres_days <= 0:
                        pres_days = 1

                    purchase_item = string_utils.xstr(prescript_row['MedicineName'])
                    purchase_item_fee = number_utils.get_integer(prescript_row['Amount']) * pres_days
                    massage_item = ''
                    massage_item_fee = 0
                    total_purchase_item_fee += purchase_item_fee
                else:
                    purchase_item = ''
                    purchase_item_fee = 0
                    massage_item = string_utils.xstr(prescript_row['MedicineName'])
                    massage_item_fee = number_utils.get_integer(prescript_row['Amount'])
                    total_massage_item_fee += massage_item_fee

                row[12] = purchase_item
                row[13] = purchase_item_fee
                row[15] = massage_item
                row[16] = massage_item_fee

                if prescript_row_no > 0:
                    row[3] = ''
                    row[7] = ''
                    row[8] = ''
                    row[9] = 0
                    row[10] = 0
                    row[11] = 0

                sheet.append(row)

        if (case_date == next_case_date and period != next_period) or (case_date != next_case_date):
            subtotal = total_regist_fee + total_diag_share_fee + total_drug_share_fee + \
                       total_purchase_item_fee + total_massage_item_fee + total_deposit_fee
            subtotal_row = [
                row[0], row[1], None,
                ins_count,
                None, None, None, None, None,
                total_regist_fee,
                total_diag_share_fee,
                total_drug_share_fee,
                None,
                total_purchase_item_fee,
                None, None,
                total_massage_item_fee,
                None,
                total_deposit_fee,
                None,
                subtotal,
            ]
            sheet.append(subtotal_row)

            for col in row_range:
                cell = sheet['{col}{row_no}'.format(col=col, row_no=sheet._current_row)]
                cell.fill = salmon
                cell.font = bold

            daily_ins_count += ins_count
            daily_regist_fee += total_regist_fee
            daily_diag_share_fee += total_diag_share_fee
            daily_drug_share_fee += total_drug_share_fee
            daily_purchase_item_fee += total_purchase_item_fee
            daily_massage_item_fee += total_massage_item_fee
            daily_deposit_fee += total_deposit_fee
            daily_total += subtotal

            ins_count = 0
            total_regist_fee = 0
            total_diag_share_fee = 0
            total_drug_share_fee = 0
            total_purchase_item_fee = 0
            total_massage_item_fee = 0
            total_deposit_fee = 0

        if case_date != next_case_date:
            sheet.append([
                row[0], '全', None,
                daily_ins_count,
                None, None, None, None, None,
                daily_regist_fee,
                daily_diag_share_fee,
                daily_drug_share_fee,
                None,
                daily_purchase_item_fee,
                None, None,
                daily_massage_item_fee,
                None,
                daily_deposit_fee,
                None,
                daily_total,
            ])
            for col in row_range:
                cell = sheet['{col}{row_no}'.format(col=col, row_no=sheet._current_row)]
                cell.fill = purple
                cell.font = bold

            daily_ins_count = 0
            daily_regist_fee = 0
            daily_diag_share_fee = 0
            daily_drug_share_fee = 0
            daily_purchase_item_fee = 0
            daily_massage_item_fee = 0
            daily_deposit_fee = 0
            daily_total = 0

    for row_no in ['E2', 'F2', 'G2', 'H2', 'I2']:
        sheet[row_no].fill = light_blue
        sheet[row_no].font = bold
        sheet[row_no].alignment = align_center

    for row_no in ['J2', 'K2', 'L2']:
        sheet[row_no].fill = light_yellow
        sheet[row_no].font = bold
        sheet[row_no].alignment = align_center

    for row_no in ['M2', 'N2', 'O2', 'P2', 'Q2', 'R2']:
        sheet[row_no].fill = light_gray
        sheet[row_no].font = bold
        sheet[row_no].alignment = align_center

    for row_no in range(2, len(sheet['A'])+1):
        for col in row_range:
            cell = sheet['{col}{row_no}'.format(col=col, row_no=row_no)]
            cell.border = border
            cell.alignment = align_center
            if row_no == 2:
                cell.font = bold

    for row_no in [2, 3]:
        for col in row_range+['V']:
            cell = sheet['{col}{row_no}'.format(col=col, row_no=row_no)]
            sheet.freeze_panes = cell

    workbook.save(excel_file_name)
    subprocess.Popen([excel_file_name], shell=True)


# 匯出巡迴醫療日報表
def export_tour_daily_list_to_excel(
        database, system_settings, excel_file_name,
        apply_date, apply_year, apply_month, apply_type_code, period, clinic_id):
    sql = '''
        SELECT * FROM insapply
        WHERE
            ApplyDate = "{apply_date}" AND
            ApplyType = "{apply_type}" AND
            ApplyPeriod = "{period}" AND
            ClinicID = "{clinic_id}" AND
            CaseType = "25"
        GROUP BY CaseDate
    '''.format(
        apply_date=apply_date,
        apply_type=apply_type_code,
        period=period,
        clinic_id=clinic_id,
    )
    rows = database.select_record(sql)
    tour_apply_count = len(rows)
    if tour_apply_count <= 0:
        return

    workbook = Workbook()
    sheet = workbook.active
    workbook.remove_sheet(sheet)
    for row in rows:
        add_tour_daily_list_sheet(
            database, system_settings, row, workbook, apply_year, apply_month,
            apply_type_code, period, clinic_id,
        )

    workbook.save(excel_file_name)
    subprocess.Popen([excel_file_name], shell=True)


def add_tour_daily_list_sheet(
        database, system_settings, row, workbook,
        apply_year, apply_month, apply_type_code, period, clinic_id):
    apply_date = nhi_utils.get_apply_date(apply_year, apply_month)
    case_date = date_utils.west_date_to_nhi_date(row['CaseDate'], '-')
    title = '{0}巡迴醫療門診日報表資料'.format(case_date)
    sheet = workbook.create_sheet(title)

    branch = '中保會{0}分會'.format(system_settings.field('健保業務').split('業務組')[0])
    sql = '''
        SELECT TourArea FROM cases
        WHERE
            CaseKey = {case_key}
    '''.format(
        case_key=row['CaseKey1'],
    )
    rows = database.select_record(sql)
    if len(rows) > 0:
        tour_area = string_utils.xstr(rows[0]['TourArea'])
    else:
        tour_area = ''

    sheet.append([
        '{apply_year}年度全民健康保險中醫門診總額醫療資源不足地區醫療服務門診日報表'.format(apply_year=apply_year-1911),
        None, None, None, None, None, None, None, None, None,
        None, None, None, None, None, None, None, None, None, None,
        None, None, None, None, None, None, None,
        '所　屬　分　會', None, branch,
    ])
    sheet.append([
        None, None, None, None, None, None, None, None, None, None,
        None, None, None, None, None, None, None, None, None, None,
        None, None, None, None, None, None, None,
        '承　辦　單　位', None,
    ])
    sheet.append([
        None, None, None, None, None, None, None, None, None, None,
        None, None, None, None, None, None, None, None, None, None,
        None, None, None, None, None, None, None,
        '醫事服務機構代碼', None, clinic_id,
    ])
    sheet.append([
        None, None, None, None, None, None, None, None, None, None,
        None, None, None, None, None, None, None, None, None, None,
        None, None, None, None, None, None, None,
        '地　　　　　點', None, tour_area,
    ])
    sheet.append([
        None, None, None, None, None, None, None, None, None, None,
        None, None, None, None, None, None, None, None, None, None,
        None, None, None, None, None, None, None,
        '核　准　代　碼', None,
    ])
    merge_cell = [
        'A1:AA5', 'AB1:AC1', 'AB2:AC2', 'AB3:AC3', 'AB4:AC4', 'AB5:AC5',
        'AD1:AF1', 'AD2:AF2', 'AD3:AF3', 'AD4:AF4', 'AD5:AF5',
    ]
    for cell in merge_cell:
        sheet.merge_cells(cell)

    for i in range(1, 6):
        row_property = sheet.row_dimensions[i]
        row_property.alignment = align_center
        row_property.font = bold

    header_row = [
        '日期', case_date, None, '時間', '08:00 - 18:00', None,
    ]
    sheet.append(header_row)
    sheet.merge_cells('B6:C6')
    sheet.merge_cells('E6:F6')
    row_property = sheet.row_dimensions[6]
    row_property.font = bold
    row_property.alignment = align_center

    '''
        'B41', 'B42', 'B43', 'B44', 'B45', 'B46',
        'B53', 'B54', 'B55', 'B56', 'B57',
        'B61', 'B62', 'B63',
    '''
    header_row = [
        '編號', '姓名', '身份證統一編號', '出生年月日', '性別', '住址', '電話',
        '診察費', '藥費(天)', '調劑費', None,
        '治療處置', None, None, None, None, None, None, None, None, None, None, None, None, None,
        '當地居民', None,
        '醫療費用', '部份負擔', '申請費用', '身份別', '備註',
    ]
    sheet.append(header_row)
    row_property.font = bold
    sheet.merge_cells('J7:K7')
    sheet.merge_cells('L7:Y7')
    sheet.merge_cells('Z7:AA7')
    row_property.alignment = align_center

    header_row = [
        None, None, None, None, None, None, None,
        None, None, 'A31', 'A32',
        'B41', 'B42', 'B43', 'B44', 'B45', 'B46',
        'B53', 'B54', 'B55', 'B56', 'B57',
        'B61', 'B62', 'B63',
        '是', '否',
    ]
    sheet.append(header_row)
    merge_cell = [
        'A7:A8', 'B7:B8', 'C7:C8', 'D7:D8', 'E7:E8', 'F7:F8', 'G7:G8',
        'H7:H8', 'I7:I8', 'AB7:AB8', 'AC7:AC8', 'AD7:AD8', 'AE7:AE8', 'AF7:AF8',
    ]
    for cell in merge_cell:
        sheet.merge_cells(cell)

    row_property = sheet.row_dimensions[7]
    row_property.alignment = align_center
    row_property.font = bold
    row_property = sheet.row_dimensions[8]
    row_property.alignment = align_center
    row_property.font = bold

    sheet.column_dimensions['B'].width = 10
    sheet.column_dimensions['C'].width = 13
    sheet.column_dimensions['D'].width = 13
    sheet.column_dimensions['F'].width = 35
    sheet.column_dimensions['G'].width = 15
    adjust_column = [
        'E',
        'J', 'K', 'L', 'M', 'N', 'O', 'P', 'Q', 'R', 'S', 'T', 'U', 'V', 'W', 'X', 'Y',
        'Z', 'AA'
    ]
    for cell in adjust_column:
        sheet.column_dimensions[cell].width = 5

    sql = '''
        SELECT *, 
               cases.TourArea,
               patient.Gender, patient.Address, patient.Telephone
        FROM insapply
            LEFT JOIN cases ON insapply.CaseKey1 = cases.CaseKey
            LEFT JOIN patient ON insapply.PatientKey = patient.PatientKey
        WHERE
            ApplyDate = "{apply_date}" AND
            insapply.ApplyType = "{apply_type}" AND
            ApplyPeriod = "{period}" AND
            ClinicID = "{clinic_id}" AND
            CaseType = "25" AND
            insapply.CaseDate = "{case_date}"
        ORDER BY Sequence
    '''.format(
        apply_date=apply_date,
        apply_type=apply_type_code,
        period=period,
        clinic_id=clinic_id,
        case_date=row['CaseDate'],
    )
    rows = database.select_record(sql)

    for row_no, row in zip(range(len(rows)), rows):
        if string_utils.xstr(row['Gender']) == '男':
            gender_code = '1'
        elif string_utils.xstr(row['Gender']) == '女':
            gender_code = '0'
        else:
            gender_code = ''

        pres_days = row['PresDays']
        pharmacy_code = nhi_utils.extract_pharmacy_code(string_utils.xstr(row['PharmacyCode']))
        pharmacy_list = [None, None]
        if pharmacy_code != '':
            pharmacy_code_dict = {
                'A31': 0, 'A32': 1,
            }
            pharmacy_list[pharmacy_code_dict[pharmacy_code]] = 'V'

        treat_list = [
            None, None, None, None, None, None, None, None, None, None, None, None, None, None,
        ]
        treat_code = string_utils.xstr(row['TreatCode1'])
        if treat_code != '':
            treat_code_dict = {
                'B41': 0, 'B42': 1, 'B43': 2, 'B44': 3, 'B45': 4, 'B46': 5,
                'B53': 6, 'B54': 7, 'B55': 8, 'B56': 9, 'B57': 10,
                'B61': 11, 'B62': 12, 'B63': 13,
            }
            treat_list[treat_code_dict[treat_code]] = 'V'

        address = string_utils.xstr(row['Address'])
        native_list = [
            None, None,
        ]
        if address == '' or tour_area in address:
            native_list[0] = 'V'
        else:
            native_list[1] = 'V'

        share_code = string_utils.xstr(row['ShareCode'])
        if share_code in ['S10', 'S20']:
            share_code = ''

        data = [
            row_no+1,
            string_utils.xstr(row['Name']),
            string_utils.xstr(row['ID']),
            string_utils.xstr(date_utils.west_date_to_nhi_date(row['Birthday'], '-')),
            gender_code,
            address,
            string_utils.xstr(row['Telephone']),
            string_utils.xstr(row['DiagCode']),
            pres_days, pharmacy_list[0], pharmacy_list[1],
            treat_list[0], treat_list[1], treat_list[2], treat_list[3], treat_list[4], treat_list[5],
            treat_list[6], treat_list[7], treat_list[8], treat_list[9], treat_list[10],
            treat_list[11], treat_list[12], treat_list[13],
            native_list[0], native_list[1],
            string_utils.xstr(row['InsTotalFee']),
            string_utils.xstr(row['ShareFee']),
            string_utils.xstr(row['InsApplyFee']),
            share_code,
        ]
        sheet.append(data)
        row_property = sheet.row_dimensions[row_no+9]
        row_property.alignment = align_center


